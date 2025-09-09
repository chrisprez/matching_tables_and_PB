import warnings
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

warnings.simplefilter("ignore", UserWarning)

try:
    # Paths and sheets
    inventory_path = '<<TARGET_FILE_PATH_FOR_INVENTORY>>'  # Put the target inventory filename here
    inventory_sheet = '<<TARGET_SHEET_NAME_FOR_INVENTORY>>'  # Put the target inventory sheet name here
    results_path = '<<TARGET_FILE_PATH_FOR_RESULTS>>'  # Put the target results filename here
    results_sheet = '<<TARGET_SHEET_NAME_FOR_RESULTS>>'  # Put the target results sheet name here
    
    def clean_value(x):
        if not isinstance(x, str):
            return x
        x = x.strip().lower()
        ip_match = re.match(r'^(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', x)
        if ip_match:
            return ip_match.group(1)
        else:
            return x.split('.')[0]
    
    # Load and clean inventory data
    # Columns D and E are chosen as an example here; adjust columns according to your requirement
    df_inventory = pd.read_excel(inventory_path, sheet_name=inventory_sheet, usecols='D:E') # Columns renamed as D and E for processing; adjust if necessary
    df_inventory.columns = ['D', 'E']  
    df_inventory = df_inventory.apply(lambda col: col.map(clean_value))
    df_inventory.dropna(how='all', inplace=True)
    df_inventory.drop_duplicates(inplace=True)
    
    # Detect duplicates in column D
    duplicates_d = df_inventory['D'].duplicated(keep=False)
    duplicated_values_d = set(df_inventory.loc[duplicates_d, 'D'])
    
    # Load and clean results data (single column 'A' used as example)
    df_results = pd.read_excel(results_path, sheet_name=results_sheet, usecols='A', header=None, names=['ColA'])
    df_results = df_results['ColA'].map(clean_value).dropna().drop_duplicates()
    result_values = set(df_results.tolist())
    
    # Open the inventory workbook for editing
    wb = load_workbook(inventory_path)
    ws = wb[inventory_sheet]
    
    # Define formats to apply
    blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
    calibri_11_font = Font(name='Calibri', size=11)
    
    # Mark columns D and E and apply Calibri font
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):  # Columns 4=D, 5=E
        row_index = row[0].row
        for col_index, cell in enumerate(row, start=4):
            cell_value = cell.value
            if isinstance(cell_value, str):
                cell_value = clean_value(cell_value)
            cell.font = calibri_11_font
            if cell_value in result_values or (col_index == 4 and cell_value in duplicated_values_d):
                cell.fill = blue_fill
    
    # Find column T (1-based) by header "Executed"
    # Comment: The column with header "Executed" is searched dynamically. Adjust header name if needed.
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    col_t_index = None
    for cell in header_row:
        if cell.value == 'Executed':
            col_t_index = cell.column
            break
    if col_t_index is None:
        raise ValueError('Column with header "Executed" not found')
    
    # Modify column T only for rows that have marked cells (in columns D or E)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_index = row[0].row
        marked = False
        for col in [4, 5]:
            cell = ws.cell(row=row_index, column=col)
            if cell.fill == blue_fill:
                marked = True
                break
        if marked:
            cell_t = ws.cell(row=row_index, column=col_t_index)
            cell_t.value = 'Executed'  
            cell_t.font = calibri_11_font
    
    # Save changes
    wb.save(inventory_path)
    print(f'File "{inventory_path}" modified successfully')
    input("Press Enter to close...")
    
except Exception as e:
    print(f'An error occurred: {e}')
    input("Press Enter to close...")
